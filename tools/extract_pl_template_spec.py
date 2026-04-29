"""Extract a deterministic JSON spec from the client P&L Excel template.

Output: assets/templates/pl_template_spec.json — consumed at runtime by the
Dart `_exportPLTemplate` to render a pixel-equivalent workbook.

Run:
    tools/.venv/bin/python tools/extract_pl_template_spec.py
"""
from __future__ import annotations

import json
import os
import sys
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter


SRC = "assets/templates/profit_and_loss_template.xlsx"
OUT = "assets/templates/pl_template_spec.json"
SHEET_NAME = "Sheet1"  # the visible P&L sheet


def color_to_hex(c) -> Optional[str]:
    """openpyxl Color → 'AARRGGBB' uppercase or None."""
    if c is None:
        return None
    if c.type == "rgb" and c.rgb:
        s = c.rgb if isinstance(c.rgb, str) else None
        if s:
            return s.upper()
    if c.type == "theme":
        return f"THEME:{c.theme}:{c.tint or 0}"
    if c.type == "indexed":
        return f"INDEXED:{c.indexed}"
    return None


def font_dict(font) -> Dict[str, Any]:
    return {
        "name": font.name,
        "size": float(font.size) if font.size else None,
        "bold": bool(font.bold),
        "italic": bool(font.italic),
        "underline": font.underline,
        "color": color_to_hex(font.color),
    }


def fill_dict(fill) -> Optional[Dict[str, Any]]:
    if fill is None or fill.patternType in (None, "none"):
        return None
    return {
        "pattern": fill.patternType,
        "fg": color_to_hex(fill.fgColor),
        "bg": color_to_hex(fill.bgColor),
    }


def alignment_dict(a) -> Dict[str, Any]:
    return {
        "h": a.horizontal,
        "v": a.vertical,
        "wrap": bool(a.wrap_text),
        "indent": a.indent or 0,
    }


def border_side(side) -> Optional[Dict[str, Any]]:
    if side is None or side.style is None:
        return None
    return {"style": side.style, "color": color_to_hex(side.color)}


def border_dict(b) -> Dict[str, Any]:
    return {
        "left": border_side(b.left),
        "right": border_side(b.right),
        "top": border_side(b.top),
        "bottom": border_side(b.bottom),
    }


def cell_spec(cell: Cell) -> Dict[str, Any]:
    val = cell.value
    if val is None:
        text = ""
    elif isinstance(val, (int, float)):
        text = val
    else:
        text = str(val)
    return {
        "ref": cell.coordinate,
        "row": cell.row,
        "col": cell.column,
        "letter": get_column_letter(cell.column),
        "value": text,
        "value_type": "number" if isinstance(text, (int, float)) else "text",
        "number_format": cell.number_format,
        "font": font_dict(cell.font),
        "fill": fill_dict(cell.fill),
        "alignment": alignment_dict(cell.alignment),
        "border": border_dict(cell.border),
    }


def main() -> int:
    wb = load_workbook(SRC, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Available sheets: {wb.sheetnames}")
        return 1
    ws = wb[SHEET_NAME]

    cols: List[Dict[str, Any]] = []
    for letter, dim in ws.column_dimensions.items():
        cols.append(
            {
                "letter": letter,
                "min": dim.min,
                "max": dim.max,
                "width": dim.width,
                "hidden": bool(dim.hidden),
                "outline": dim.outlineLevel or 0,
            }
        )
    cols.sort(key=lambda x: (x["min"] or 0))

    rows: List[Dict[str, Any]] = []
    for idx, dim in ws.row_dimensions.items():
        rows.append(
            {
                "index": idx,
                "height": dim.height,
                "hidden": bool(dim.hidden),
                "outline": dim.outlineLevel or 0,
            }
        )
    rows.sort(key=lambda x: x["index"])

    merges = [str(r) for r in ws.merged_cells.ranges]

    cells: List[Dict[str, Any]] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None and cell.has_style is False:
                continue
            cells.append(cell_spec(cell))

    spec = {
        "source": SRC,
        "sheet": SHEET_NAME,
        "dimensions": ws.dimensions,
        "default_row_height": ws.sheet_format.defaultRowHeight,
        "default_col_width": ws.sheet_format.defaultColWidth,
        "columns": cols,
        "rows": rows,
        "merges": merges,
        "cells": cells,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(spec, f, indent=2, default=str)
    print(f"Wrote {OUT} ({len(cells)} cells, {len(merges)} merges, {len(cols)} cols, {len(rows)} rows)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
